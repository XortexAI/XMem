from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

NAVY = "1B3A6B"
GOLD = "C9A84C"
WHITE = "FFFFFF"
LGRAY = "F4F6FA"
MGRAY = "8A99B0"
DGRAY = "2D3748"
TEAL = "0D9488"
RED = "C0392B"
GREEN_TX = "007A3D"
BLUE_TX = "0000FF"

INR = '₹#,##0;(₹#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'

def side(style="thin", color="D1D5DB"):
    return Side(style=style, color=color)

def border():
    s = side()
    return Border(top=s, bottom=s, left=s, right=s)

def cell(ws, row, col, val=None, bold=False, bg=WHITE, fg=BLACK_TX,
         size=10, align="left", fmt=None, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, color=fg, size=size, italic=italic)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border()
    if fmt:
        c.number_format = fmt
    return c

def merge(ws, r1, c1, r2, c2, val=None, bold=False, bg=WHITE, fg=BLACK_TX):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=val)
    c.font = Font(name="Arial", bold=bold, color=fg, size=11)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

# ================= MASTER SHEET =================
ws = wb.create_sheet("Master Budget")

headers = ["Item", "Amount", "%"]
for i, h in enumerate(headers, 1):
    cell(ws, 1, i, h, bold=True, bg=NAVY, fg=WHITE)

data = [
    ("1st Prize", 25000),
    ("2nd Prize", 12000),
    ("3rd Prize", 6000),
    ("Marketing", 78000),
    ("Team", 25000),
    ("Misc", 25000),
]

row = 2
for label, amount in data:
    cell(ws, row, 1, label)
    cell(ws, row, 2, amount, fg=BLUE_TX, align="center", fmt=INR)
    row += 1

# Total
cell(ws, row, 1, "TOTAL", bold=True)
cell(ws, row, 2, f"=SUM(B2:B{row-1})", bold=True, fmt=INR)

# Percent column
for r in range(2, row):
    cell(ws, r, 3, f"=B{r}/B{row}", fmt=PCT, align="center")

# ================= PRIZE SHEET =================
ws2 = wb.create_sheet("Prize Structure")

headers = ["Rank", "Cash Prize"]
for i, h in enumerate(headers, 1):
    cell(ws2, 1, i, h, bold=True, bg=NAVY, fg=WHITE)

prizes = [
    ("1st", 25000),
    ("2nd", 12000),
    ("3rd", 6000),
]

for i, (rank, val) in enumerate(prizes, start=2):
    cell(ws2, i, 1, rank)
    cell(ws2, i, 2, val, fmt=INR, align="center")

# ================= SAVE FILE =================
wb.save("DSA_Budget.xlsx")

print("✅ Excel file saved as DSA_Budget.xlsx in your folder")